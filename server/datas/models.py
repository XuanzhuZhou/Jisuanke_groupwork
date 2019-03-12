# -*- coding: utf-8 -*-

'''本文件用于建立各种记录的模型'''
import csv
import os
from datetime import datetime

from django.db import models
from django.utils import timezone

import xlrd
from courses.models import Courses, Section, Teacher, get_teacher
from courses.views import init_sections
from customers.models import Customer
from openpyxl import load_workbook
from statuscode import DATE_DECODE, STATUS_CODE
from users.models import SellerInfo, User

# Create your models here.


class OtherResource(models.Model):
    '''第三方来源'''
    customer = models.ForeignKey(Customer, on_delete=models.CASCADE)
    resource = models.CharField(max_length=200)
    resource_date = models.DateField()


def get_section():
    '''返回一个标记为已删除的Section行'''
    return Section.objects.get_or_create(name='deleted')[0]


class AuditionRecord(models.Model):
    '''每次试听记录'''
    customer = models.ForeignKey(
        Customer, related_name='audition_record', on_delete=models.CASCADE)
    section = models.ForeignKey(
        Section, related_name='audition_record',
        null=True, on_delete=models.SET(get_section)
    )
    is_signedup = models.BooleanField(default=False, blank=False)
    date = models.DateTimeField(default=timezone.now)
    info = models.TextField(null=True, blank=True)


class SellRecord(models.Model):
    '''注册记录'''
    customer = models.ForeignKey(
        Customer, related_name='register_record', on_delete=models.CASCADE)
    seller = models.ForeignKey(
        SellerInfo, null=True,
        related_name='sell_record', on_delete=models.SET_NULL)
    money = models.FloatField()
    date = models.DateTimeField(default=timezone.now)
    trade_no = models.CharField(max_length=20)


def get_user():
    '''返回一个标记为已删除的User行'''
    return User.objects.get_or_create(username='deleted')[0]


def get_customer():
    '''返回一个标记为已删除的 customer 行'''
    return Customer.objects.get_or_create(
        user=get_user(), child_name='deleted')[0]


def get_course():
    '''返回一个标记为已删除的 course 行'''
    return Courses.objects.get_or_create(
        teacher_id=get_teacher(), name='deleted')[0]


class RefundRecord(models.Model):
    '''退费记录'''
    customer = models.ForeignKey(
        Customer, related_name='refund_record',
        on_delete=models.SET(get_customer))
    course = models.ForeignKey(
        Courses, related_name='refund_record',
        on_delete=models.SET(get_course))
    refund = models.FloatField(null=True, blank=True)
    cc = models.ForeignKey(
        User, related_name='review_refund', null=True,
        on_delete=models.SET(get_user))
    date = models.DateTimeField(default=timezone.now)
    deal_date = models.DateTimeField(null=True, blank=True)
    is_passed = models.BooleanField(default=False)
    is_paid = models.BooleanField(default=False)
    reason = models.TextField()


class CcRecord(models.Model):
    '''常规课报名记录'''
    customer = models.ForeignKey(
        Customer, related_name='signup_record', on_delete=models.CASCADE)
    cc = models.ForeignKey(User, related_name='sell_record',
                           null=True, on_delete=models.SET_NULL)
    money = models.FloatField()
    date = models.DateTimeField(default=timezone.now)
    info = models.TextField()
    is_paid = models.BooleanField(default=False)
    trade_no = models.CharField(max_length=20)
    course = models.ForeignKey(Courses, related_name='signup_record',
                               null=True, on_delete=models.SET(get_course))


class OperationLogs(models.Model):
    '''操作日志'''
    TYPE_CHOICES = (
        (0, 'refund'),
        (1, 'del_user'),
        (2, 'del_customer'),
        (3, 'add_user'),
    )

    operator = models.ForeignKey(
        User, related_name='user_log', on_delete=models.SET(get_user))
    type = models.PositiveSmallIntegerField(
        choices=TYPE_CHOICES, default=0)
    username = models.CharField(max_length=20)
    date = models.DateTimeField(default=timezone.now)
    info = models.TextField(max_length=500)


def xlsx_import(file):
    '''
    本函数用来导入xlsx文件

    :传入数据: 文件 file

    :传出数据: colnames(文件数据名字列表)  the_list(数据列表)
    '''
    try:
        data = load_workbook(file)
    except IOError:
        return (STATUS_CODE['Unable To Import File'], [])
    sheets = data.sheetnames
    table = data[sheets[0]]
    row_num = table.max_row
    colnames = []
    for i in table['1']:
        colnames.append(i.value)
    the_list = []
    for num in range(2, row_num+1):
        row = []
        for i in table[str(num)]:
            row.append(i.value)
        the_list.append(row)
    return (colnames, the_list)


def csv_import(file):
    '''
    本函数用来导入csv文件

    :传入数据: 文件 file

    :传出数据: colnames(文件数据名字列表)  the_list(数据列表)
    '''
    colnames = []
    the_list = []
    with open(file, 'r') as out_file:
        reader = csv.reader(out_file)
        for item in reader:
            if reader.line_num == 1:
                colnames = item
            else:
                the_list.append(item)
    return (colnames, the_list)


def xls_import(file, colnameidex=0):
    '''
    本函数用来导入xls文件

    :传入数据: 文件 file

    :传出数据: colnames(文件数据名字列表)  the_list(数据列表)
    '''
    try:
        data = xlrd.open_workbook(file)
    except IOError:
        return (STATUS_CODE['Unable To Import File'], [])
    table = data.sheet_by_index(0)
    row_num = table.nrows
    colnames = table.row_values(colnameidex)
    the_list = []
    for num in range(1, row_num):
        row = table.row_values(num)
        if row:
            app = []
            for i in range(len(colnames)):
                app.append(row[i])
            the_list.append(app)
    return (colnames, the_list)


def file_import_func(file_name):
    '''
    本函数用来根据文件类型来返回处理函数

    :传入数据: 文件名 file_name

    :传出数据: 函数名
    '''
    if os.path.splitext(file_name)[-1] == ".xlsx":
        return xlsx_import
    if os.path.splitext(file_name)[-1] == ".xls":
        return xls_import
    if os.path.splitext(file_name)[-1] == ".csv":
        return csv_import
    return STATUS_CODE['Error File Type']


def customer_file_seek(colnames):
    '''本函数用来找到家长文件中相应列'''
    try:
        cols = {}
        cols['child_name_col'] = colnames.index('学生姓名')
        cols['gender_col'] = colnames.index('性别')
        cols['parent_col'] = colnames.index('家长')
        cols['tel_col'] = colnames.index('联系电话')
        cols['date_col'] = colnames.index('流量获得日期')
        cols['resource_col'] = colnames.index('流量来源')
        cols['classin_id_col'] = colnames.index('Classin账号')
        cols['Classin_name_col'] = colnames.index('Classin昵称')
        cols['cc_name_col'] = colnames.index('服务课顾')
    except ValueError:
        return STATUS_CODE['File Data Error']
    return cols


def import_customer_info(file, file_name, superuser):
    '''
    本函数用来导入客户信息

    :传入数据: 文件 file 文件名 file_name 用户对象 superuser

    :传出数据: 状态码 0或703或702或701或 和错误列表
    '''
    file_import = file_import_func(file_name)
    if file_import is not STATUS_CODE['Error File Type']:
        (colnames, data) = file_import(file)
        if colnames == STATUS_CODE['Unable To Import File']:
            return (STATUS_CODE['Unable To Import File'], [])
    else:
        return (STATUS_CODE['Error File Type'], [])
    cols = customer_file_seek(colnames)
    if not isinstance(cols, dict):
        return (cols, [])
    errors = []
    for row in data:
        info = {}
        for key in cols:
            info[key] = row[cols[key]]
        error = add_customer(info, superuser, file_name)
        if error is not STATUS_CODE['Success']:
            errors.append(error)
    return (STATUS_CODE['Success'], errors)


def add_customer(info, superuser, file_name):
    '''
    本函数用来依据数据自带你导入客户数据

    :传入数据: 数据字典 info 文件名 file_name 用户对象 superuser

    :传出数据: 状态码 0或705  错误则包含错误明细
    '''
    username = info['tel_col']
    if not username.isdigit() or not len(username) == 11:
        return {'msg': STATUS_CODE['Frontend Value Error'],
                'name': username}
    user = User.objects.get_or_create(username=username, user_type=1)
    if user[1]:
        user[0].set_password(username[6:])
        OperationLogs.objects.create(
            operator=superuser, type=3,
            username=username, info='Import from file' + file_name
        ).save()
    user[0].save()
    try:
        child_name = info['child_name_col']
        cc_name = info['cc_name_col']
        cc = User.objects.get_or_create(username=cc_name)[0]
        customer = Customer.objects.get_or_create(
            user=user[0], child_name=child_name)[0]
        customer.cc = cc
        customer.parent_name = info['parent_col']
        customer.classin_id = info['classin_id_col']
        customer.classin_name = info['Classin_name_col']
        customer.is_paid = True
        customer.gender = info['gender_col']
        customer.save()
        OtherResource.objects.create(customer=customer,
                                     resource=info['resource_col'],
                                     resource_date=info['date_col']).save()
    except (ValueError, TypeError):
        return {'msg': STATUS_CODE['Frontend Value Error'],
                'name': username}
    return STATUS_CODE['Success']


def course_file_seek(colnames):
    '''本函数用来找到课程文件中相应列'''
    try:
        cols = {}
        cols['name'] = colnames.index('课程名')
        cols['teacher_name'] = colnames.index('教师名')
        cols['teacher_tel'] = colnames.index('教师电话')
        cols['weekday'] = colnames.index('星期')
        cols['time'] = colnames.index('上课时间')
        cols['total_sec'] = colnames.index('总课时')
        cols['course_length'] = colnames.index('课长')
        cols['price'] = colnames.index('价格')
        cols['date_start'] = colnames.index('开始日期')
    except ValueError:
        return STATUS_CODE['File Data Error']
    return cols


def add_course(info):
    '''本函数根据数据字典用来添加课程'''
    teacher = Teacher.objects.filter(
        name=info['teacher_name'], tel=info['teacher_tel'])
    if teacher.count() == 0:
        error = {'msg': STATUS_CODE['Teacher Not Found'],
                 'name': info['teacher_name'] + info['teacher_tel']}
        return error
    course, created = Courses.objects.get_or_create(
        name=info['name'], teacher_id=teacher[0])
    if not created:
        error = {'msg': STATUS_CODE['Course Already Exists'],
                 'name': info['name']}
        return error
    try:
        key = list(DATE_DECODE.values()).index(info['weekday'])
        course.date = key + 1
    except (ValueError, TypeError):
        error = {'msg': STATUS_CODE['Frontend Value Error'],
                 'name': info['name']}
        course.delete()
        return error
    course.total_sec = int(info['total_sec'])
    course.price = int(info['price'])
    course.course_length = int(info['course_length'])
    course.save()
    msg = add_section(info['time'], info['date_start'], course)
    if msg != STATUS_CODE['Success']:
        course.delete()
    return msg


def add_section(start_time, start_date, course):
    '''本函数用来根据导入文件课程添加课节'''

    if isinstance(start_time, str) and isinstance(start_date, str):
        try:
            course.time = datetime.strftime(start_time, '%H:%M:%S')
            start_date = datetime.strftime(start_date, '%Y-%m-%d').date()
        except (TypeError, TypeError):
            error = {'msg': STATUS_CODE['Error Time Range'],
                     'name': course.name}
            return error
    else:
        try:
            course.time = start_time
            start_date = start_date.date()
        except (ValueError, TypeError):
            error = {'msg': STATUS_CODE['Error Time Range'],
                     'name': course.name}
            return error
    course.save()
    try:
        init_sections(course, start_date)
    except (ValueError, TypeError):
        error = {'msg': STATUS_CODE['Error Time Range'],
                 'name': course.name}
        return error
    return STATUS_CODE['Success']


def import_courses(file, file_name):
    '''
    本函数用来从文件中导入客户信息

    :传入数据: 文件 file 文件名 file_name

    :传出数据: 状态码 0或703或702或701或 和错误列表

    '''
    file_import = file_import_func(file_name)
    if file_import is not STATUS_CODE['Error File Type']:
        (colnames, data) = file_import(file)
        if colnames == STATUS_CODE['Unable To Import File']:
            return (STATUS_CODE['Unable To Import File'], [])
    else:
        return (STATUS_CODE['Error File Type'], [])
    cols = course_file_seek(colnames)
    if not isinstance(cols, dict):
        return (cols, [])
    errors = []
    for row in data:
        info = {}
        for key in cols:
            info[key] = row[cols[key]]
        error = add_course(info)
        if error is not STATUS_CODE['Success']:
            errors.append(error)
    return (STATUS_CODE['Success'], errors)
