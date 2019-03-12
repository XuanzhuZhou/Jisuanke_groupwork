# -*- coding: utf-8 -*-

'''本文件用于测试'''
import csv
import json
import os
from datetime import date, datetime, timedelta

from django.test import Client, TestCase

import xlwt
from courses.models import Courses, Section, Teacher
from customers.models import Customer
from openpyxl import Workbook
from statuscode import STATUS_CODE, USER_TYPE, DATE_CODE
from users.models import SellerInfo, User

from .models import (CcRecord, AuditionRecord, SellRecord,
                     get_customer, import_courses, import_customer_info)


class SellerGradesTestCase(TestCase):
    '''本类用于测试超管获取当天试听人数、试听收入、报名人数、报名收入和地推业绩'''

    def setUp(self):
        superuser = User.objects.create(
            username='superuser', user_type=USER_TYPE['superuser'])
        superuser.set_password('superuser')
        superuser.save()
        user_cus = User.objects.create(username='15903910001')
        seller = User.objects.create(
            username='seller', user_type=USER_TYPE['seller'])
        seller2 = User.objects.create(
            username='seller2', user_type=USER_TYPE['seller'])
        seller_info = SellerInfo.objects.create(
            seller=seller, city='Beijing', price=29.9)
        seller_info2 = SellerInfo.objects.create(
            seller=seller2, city='Shanghai', price=39.9)
        for i in range(6):
            Customer.objects.create(user=user_cus, child_name='child'+str(i))
            Customer.objects.create(user=user_cus, child_name='child2'+str(i))
            AuditionRecord.objects.create(
                customer=Customer.objects.get(child_name='child'+str(i)),
                date=datetime.now()-timedelta(hours=i*24))
            AuditionRecord.objects.create(
                customer=Customer.objects.get(child_name='child2'+str(i)),
                date=datetime.now()-timedelta(hours=i*24*30))
            SellRecord.objects.create(
                customer=Customer.objects.get(child_name='child'+str(i)),
                seller=seller_info, money=29.9,
                date=datetime.now()-timedelta(hours=i*24))
            SellRecord.objects.create(
                customer=Customer.objects.get(child_name='child2'+str(i)),
                seller=seller_info2, money=39.9,
                date=datetime.now()-timedelta(hours=i*24*30))
            CcRecord.objects.create(
                customer=Customer.objects.get(child_name='child'+str(i)),
                money=2999.9, is_paid=True,
                date=datetime.now()-timedelta(hours=i*24))
            CcRecord.objects.create(
                customer=Customer.objects.get(child_name='child2'+str(i)),
                money=2999.9, is_paid=True,
                date=datetime.now()-timedelta(hours=i*24*30))

    def test_get_grades(self):
        '''测试当天有销售额'''
        client = Client()
        client.login(username='superuser', password='superuser')
        response = client.post('/data/today_grades',
                               content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(content['audition_count'] == 2)

    def test_all_sellgrades(self):
        '''测试超管获取地推当天总业绩'''
        client = Client()
        data = {
            'name': ''
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/today_seller_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(content['today_sellcount'] == 2)

    def test_sell_today_grade(self):
        '''测试超管获取当天某一个地推的业绩'''
        client = Client()
        data = {
            'name': 'seller'
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/today_seller_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(content['today_sellcount'] == 1)

    def test_sell_week_grades(self):
        '''测试超管获取最近七天地推人员业绩'''
        client = Client()
        data = {
            'name': ''
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/week_seller_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue('list' in content)

    def test_sell_week_grade(self):
        '''测试超管获取最近七天某个地推人员业绩'''
        client = Client()
        data = {
            'name': 'seller'
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/week_seller_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue('list' in content)

    def test_nosell_week_grade(self):
        '''测试超管获取最近七天某个地推人员业绩时用户名不存在'''
        client = Client()
        data = {
            'name': 'noseller'
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/week_seller_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue('error' in content['list'][0])


class SellerMonthsGradesTestCase(TestCase):
    '''本类用于测试超管获取今年各月地推人员销售情况'''

    def setUp(self):
        '''本函数用来初始化'''
        superuser = User.objects.create(
            username='superuser', user_type=USER_TYPE['superuser'])
        superuser.set_password('superuser')
        superuser.save()
        user_cus = User.objects.create(username='15903910001')
        seller = User.objects.create(
            username='seller', user_type=USER_TYPE['seller'])
        seller_info = SellerInfo.objects.create(
            seller=seller, city='Beijing', price=29.9)
        for i in range(date.today().month):
            Customer.objects.create(user=user_cus, child_name='child'+str(i))
            AuditionRecord.objects.create(
                customer=Customer.objects.get(child_name='child'+str(i)),
                date=datetime.now()-timedelta(hours=i*24*30))
            SellRecord.objects.create(
                customer=Customer.objects.get(child_name='child'+str(i)),
                seller=seller_info, money=29.9,
                date=datetime.now()-timedelta(hours=i*24*30))
            CcRecord.objects.create(
                customer=Customer.objects.get(child_name='child'+str(i)),
                money=2999.9, is_paid=True,
                date=datetime.now()-timedelta(hours=i*24*30))

    def test_sell_month_grades(self):
        '''本函数用来测试全体地推人员月度业绩'''
        client = Client()
        data = {
            'name': ''
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/months_seller_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue('list' in content)

    def test_sell_month_grade(self):
        '''本函数用来测试单个地推人员获取月度业绩'''
        client = Client()
        data = {
            'name': 'seller'
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/months_seller_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue('list' in content)

    def test_nosell_month_grade(self):
        '''本函数用来测试用户名不存在'''
        client = Client()
        data = {
            'name': 'noseller'
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/months_seller_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(content['list'][0]['error'], STATUS_CODE['Not Found'])


class SellerYearsGradesTestCase(TestCase):
    '''本类用于测试超管获取两年地推人员销售情况'''

    def setUp(self):
        '''本函数用来初始化'''
        superuser = User.objects.create(
            username='superuser', user_type=USER_TYPE['superuser'])
        superuser.set_password('superuser')
        superuser.save()
        user_cus = User.objects.create(username='15903910001')
        seller = User.objects.create(
            username='seller', user_type=USER_TYPE['seller'])
        seller_info = SellerInfo.objects.create(
            seller=seller, city='Beijing', price=29.9)
        for i in range(date.today().month):
            Customer.objects.create(user=user_cus, child_name='child'+str(i))
            AuditionRecord.objects.create(
                customer=Customer.objects.get(child_name='child'+str(i)),
                date=datetime.now()-timedelta(hours=i*24*365))
            SellRecord.objects.create(
                customer=Customer.objects.get(child_name='child'+str(i)),
                seller=seller_info, money=29.9,
                date=datetime.now()-timedelta(hours=i*24*365))
            CcRecord.objects.create(
                customer=Customer.objects.get(child_name='child'+str(i)),
                money=2999.9, is_paid=True,
                date=datetime.now()-timedelta(hours=i*24*365))

    def test_sell_year_grade(self):
        '''本函数用来测试获取seller两年业绩'''
        client = Client()
        data = {
            'name': 'seller'
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/years_seller_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue('last_list' in content)
        self.assertTrue('this_list' in content)

    def test_sell_year_grades(self):
        '''本函数用来测试获取全体seller年度业绩'''
        client = Client()
        data = {
            'name': ''
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/years_seller_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue('last_list' in content)
        self.assertTrue('this_list' in content)

    def test_unsell_year_grades(self):
        '''本函数用来测试用户名不存在的情况'''
        client = Client()
        data = {
            'name': 'noseller'
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/years_seller_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(content['last_list'][0]['error'], 1)
        self.assertEqual(content['this_list'][0]['error'], 1)


class CcGradesTestCase(TestCase):
    '''本函数用来初始化'''

    def setUp(self):
        superuser = User.objects.create(
            username='superuser', user_type=USER_TYPE['superuser'])
        superuser.set_password('superuser')
        superuser.save()
        user_cus = User.objects.create(username='15903910001')
        seller = User.objects.create(
            username='seller', user_type=USER_TYPE['seller'])
        seller_info = SellerInfo.objects.create(
            seller=seller, city='Beijing', price=29.9)
        consults = User.objects.create(
            username='cc', user_type=USER_TYPE['consultant'])
        for i in range(7):
            Customer.objects.create(
                user=user_cus, child_name='child'+str(i), cc=consults,
                date_cc=date.today()-timedelta(hours=24*i))
            AuditionRecord.objects.create(
                customer=Customer.objects.get(child_name='child'+str(i)),
                date=datetime.now()-timedelta(hours=i*24))
            SellRecord.objects.create(
                customer=Customer.objects.get(child_name='child'+str(i)),
                seller=seller_info, money=29.9,
                date=datetime.now()-timedelta(hours=i*24))
            CcRecord.objects.create(
                customer=Customer.objects.get(child_name='child'+str(i)),
                money=2999.9, is_paid=True,
                date=datetime.now()-timedelta(hours=i*24), cc=consults)

    def test_get_ccgrades(self):
        '''测试超管获取课顾当天总业绩'''
        client = Client()
        data = {
            'name': ''
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/today_cc_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(content['error'], STATUS_CODE['Success'])

    def test_get_ccgrade(self):
        '''测试超管获取某个课顾当天总业绩'''
        client = Client()
        data = {
            'name': 'cc'
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/today_cc_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(content['error'], STATUS_CODE['Success'])

    def test_get_noccgrade(self):
        '''测试输入课顾姓名不存在'''
        client = Client()
        data = {
            'name': 'nocc'
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/today_cc_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(content['error'], STATUS_CODE['User Not Found'])

    def test_get_week_ccgrades(self):
        '''测试超管获取课顾最近七天总业绩'''
        client = Client()
        data = {
            'name': ''
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/week_cc_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue('list' in content)

    def test_get_week_noccgrades(self):
        '''测试超管获取课顾最近七天总业绩'''
        client = Client()
        data = {
            'name': 'nocc'
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/week_cc_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            content['list'][0]['error'], STATUS_CODE['User Not Found'])

    def test_get_week_ccgrade(self):
        '''测试超管获取某个课顾最近七天总业绩'''
        client = Client()
        data = {
            'name': 'cc'
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/week_cc_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue('list' in content)


class CcMonthsGradesTestCase(TestCase):
    '''本类用于测试超管获取今年各月课程顾问销售情况'''

    def setUp(self):
        superuser = User.objects.create(
            username='superuser', user_type=USER_TYPE['superuser'])
        superuser.set_password('superuser')
        superuser.save()
        user_cus = User.objects.create(username='15903910001')
        seller = User.objects.create(
            username='seller', user_type=USER_TYPE['seller'])
        seller_info = SellerInfo.objects.create(
            seller=seller, city='Shanghai', price=39.9)
        cur_month = date.today().month
        consults = User.objects.create(
            username='cc', user_type=USER_TYPE['consultant'])
        for i in range(cur_month):
            Customer.objects.create(
                user=user_cus, child_name='child'+str(i), cc=consults,
                date_cc=date.today()-timedelta(hours=i*24*30))
            AuditionRecord.objects.create(
                customer=Customer.objects.get(child_name='child'+str(i)),
                date=datetime.now()-timedelta(hours=i*24*30))
            SellRecord.objects.create(
                customer=Customer.objects.get(child_name='child'+str(i)),
                seller=seller_info, money=29.9,
                date=datetime.now()-timedelta(hours=i*24*30))
            CcRecord.objects.create(
                customer=Customer.objects.get(child_name='child'+str(i)),
                money=4396, is_paid=True,
                date=datetime.now()-timedelta(hours=i*24*30), cc=consults)

    def test_cc_month_grades(self):
        '''测试超管获取全部课程顾问每个月的销售业绩'''
        client = Client()
        data = {
            'name': ''
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/months_cc_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue('list' in content)

    def test_cc_month_grade(self):
        '''测试超管获取某个课程顾问每个月的销售业绩'''
        client = Client()
        data = {
            'name': 'cc'
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/months_cc_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue('list' in content)

    def test_nocc_month_grade(self):
        '''测试课程顾问用户名不存在'''
        client = Client()
        data = {
            'name': 'nocc'
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/months_cc_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            content['list'][0]['error'], STATUS_CODE['User Not Found'])


class CcYearsGradesTestCase(TestCase):
    '''本类用于测试超管获取两年课程顾问销售情况'''

    def setUp(self):
        superuser = User.objects.create(
            username='superuser', user_type=USER_TYPE['superuser'])
        superuser.set_password('superuser')
        superuser.save()
        user_cus = User.objects.create(username='15903910001')
        seller = User.objects.create(
            username='seller', user_type=USER_TYPE['seller'])
        seller_info = SellerInfo.objects.create(
            seller=seller, city='Shanghai', price=39.9)
        consults = User.objects.create(
            username='cc', user_type=USER_TYPE['consultant'])
        for i in range(2):
            Customer.objects.create(
                user=user_cus, child_name='child'+str(i), cc=consults,
                date_cc=date.today()-timedelta(hours=i*24*365))
            AuditionRecord.objects.create(
                customer=Customer.objects.get(child_name='child'+str(i)),
                date=datetime.now()-timedelta(hours=i*24*365))
            SellRecord.objects.create(
                customer=Customer.objects.get(child_name='child'+str(i)),
                seller=seller_info, money=29.9,
                date=datetime.now()-timedelta(hours=i*24*365))
            CcRecord.objects.create(
                customer=Customer.objects.get(child_name='child'+str(i)),
                money=4369, is_paid=True,
                date=datetime.now()-timedelta(hours=i*24*365), cc=consults)

    def test_cc_year_grade(self):
        '''测试超管获取某个课程顾问的两年业绩对比情况'''
        client = Client()
        data = {
            'name': 'cc'
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/years_cc_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue('last_list' in content)
        self.assertTrue('this_list' in content)

    def test_cc_year_grades(self):
        '''测试超管获取全部课程顾问的两年业绩对比情况'''
        client = Client()
        data = {
            'name': ''
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/years_cc_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue('last_list' in content)
        self.assertTrue('this_list' in content)

    def test_nocc_year_grades(self):
        '''测试课程顾问用户名不存在'''
        client = Client()
        data = {
            'name': 'nocc'
        }
        client.login(username='superuser', password='superuser')
        response = client.post('/data/years_cc_grades',
                               data, content_type='application/json')
        content = json.loads(response.content)
        print(content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue('last_list' in content)
        self.assertTrue('this_list' in content)


class CitySellerNumsTestCase(TestCase):
    '''本类用于获取各城市的地推人数'''

    def setUp(self):
        '''本函数用来初始化'''
        superuser = User.objects.create(
            username='superuser', user_type=USER_TYPE['superuser'])
        superuser.set_password('superuser')
        superuser.save()
        seller = User.objects.create(
            username='seller', user_type=USER_TYPE['seller'])
        citys = ['北京', '上海', '广州', '深圳', '天津', '上海', '石家庄', '上海']
        for city in citys:
            SellerInfo.objects.create(seller=seller, city=city, price=39.9)

    def test_citys_num(self):
        '''本函数用来测试各城市地推人数'''
        client = Client()
        client.login(username='superuser', password='superuser')
        response = client.post('/data/get_seller_distribution',
                               content_type='application/json')
        content = json.loads(response.content)
        print(content)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(content['list']), 6)


class AuditionGradesTest(TestCase):
    '''本类用来测试查看试听课老师的试听转化率'''

    def setUp(self):
        superuser = User.objects.create(
            username='superuser', user_type=USER_TYPE['superuser'])
        superuser.set_password('superuser')
        superuser.save()
        teacher_first = Teacher.objects.create(name='王老师')
        teacher_second = Teacher.objects.create(name='李老师')
        Courses.objects.create(name="课程1", teacher_id=teacher_first)
        Courses.objects.create(name="课程3", teacher_id=teacher_first)
        Courses.objects.create(name='课程2', teacher_id=teacher_second)
        for course in Courses.objects.all():
            for i in range(10):
                section = Section.objects.create(
                    course_id=course,
                    count=i,
                    date=(datetime.strptime("2018-08-24", "%Y-%m-%d") -
                          (i-1)*timedelta(weeks=1)),
                    start_time=datetime.strptime('05:20:00', "%H:%M:%S"),
                    end_time=datetime.strptime('06:00:00', "%H:%M:%S")
                )
                for j in range(i):
                    is_signedup = (
                        True if j % 2 == 0 else False)
                    AuditionRecord.objects.create(
                        customer=get_customer(),
                        section=section,
                        is_signedup=is_signedup
                    )

    def test_audition_gradee_range(self):
        '''本函数用来测试单个老师试听转化率查询'''
        client = Client()
        client.login(username='superuser', password='superuser')
        response = client.post(
            '/data/audition_grades',
            {
                'name': '王老师',
                'date_from': '2018-07-10',
                'date_to': '2018-08-01'
            },
            content_type='application/json'
        )
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue('list' in content)

    def test_audition_grades(self):
        '''本函数用来测试单个老师试听转化率查询'''
        client = Client()
        client.login(username='superuser', password='superuser')
        response = client.post(
            '/data/audition_grades',
            {
                'name': '王老师',
                'date_from': None,
                'date_to': None
            },
            content_type='application/json'
        )
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue('week' in content)

    def test_audition_value_erorr(self):
        '''本函数用来测试单个老师试听转化率查询传入错误时'''
        client = Client()
        client.login(username='superuser', password='superuser')
        response = client.post(
            '/data/audition_grades',
            {
                'name': '王老师',
                'date_to': None
            },
            content_type='application/json'
        )
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(
            content['msg'], STATUS_CODE['Frontend Value Error'])

    def test_audition_all(self):
        '''本函数用来测试未输入名字时查询各个老师转化率请求'''
        client = Client()
        client.login(username='superuser', password='superuser')
        response = client.post(
            '/data/audition_grades',
            {
                'name': "",
                'date_from': None,
                'date_to': None
            },
            content_type='application/json'
        )
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue('list' in content)

    def test_audition_all_range(self):
        '''本函数用来测试未输入名字时按时间查询各个老师转化率请求'''
        client = Client()
        client.login(username='superuser', password='superuser')
        response = client.post(
            '/data/audition_grades',
            {
                'name': "",
                'date_from': '2018-06-10',
                'date_to': '2018-08-02'
            },
            content_type='application/json'
        )
        content = json.loads(response.content)
        self.assertEqual(response.status_code, 200)
        self.assertTrue('list' in content)


class FileImportTestCase(TestCase):
    '''本类用于文件导入数据功能'''

    def setUp(self):
        '''本函数用来初始化'''
        self.superuser = User.objects.create(
            username='superuser', user_type=USER_TYPE['superuser'])
        self.superuser.set_password('superuser')
        self.superuser.save()
        Teacher.objects.create(name="测试老师", tel='测试电话')
        self.name_list = ['学生姓名', '性别', '家长', '联系电话',
                          '流量获得日期', '流量来源', 'Classin账号',
                          'Classin昵称', '服务课顾']
        self.data_list = ['测试姓名', '男', '测试家长', '12345678910',
                          '2018-02-23', '来源信息', 'Classin账号',
                          'Classin昵称', '客服姓名']
        self.data_list2 = ['错误姓名', '男', '测试家长', '1234567890',
                           '2018-02-23', '来源信息', 'Classin账号',
                           'Classin昵称', '客服姓名']
        self.user_name = ['用户名', '性别', '初始密码', '邮箱']
        self.user_data = ['重复用户名', '男', '12345', '测试邮箱']
        self.user_data2 = ['重复用户名', '男', '12345', '测试邮箱']

    def test_check_import_xls(self):
        '''测试超级管理员导入xls数据功能'''
        new_file = xlwt.Workbook()
        sheet1 = new_file.add_sheet('Sheet')
        for i in range(len(self.name_list)):
            sheet1.write(0, i, self.name_list[i])
            sheet1.write(1, i, self.data_list[i])
            sheet1.write(2, i, self.data_list2[i])
        new_file.save('test.xls')
        result = import_customer_info('test.xls', 'test.xls', self.superuser)
        os.remove('test.xls')
        self.assertEqual(len(result[1]), 1)

    def test_check_import_xlsx(self):
        '''测试超级管理员导入xlsx数据功能'''
        new_file = Workbook()
        sheet = new_file.active
        sheet.append(self.name_list)
        sheet.append(self.data_list)
        sheet.append(self.data_list2)
        new_file.save('test.xlsx')
        result = import_customer_info(
            'test.xlsx', 'test.xlsx', self.superuser)
        os.remove('test.xlsx')
        self.assertEqual(len(result[1]), 1)

    def test_check_import_csv(self):
        '''测试超级管理员导入csv数据功能'''
        with open('test.csv', 'w') as thefile:
            writer = csv.writer(thefile)
            writer.writerow(self.name_list)
            writer.writerow(self.data_list)
            writer.writerow(self.data_list2)
        result = import_customer_info('test.csv', 'test.csv', self.superuser)
        os.remove('test.csv')
        self.assertEqual(len(result[1]), 1)

    def test_check_import_course(self):
        '''本测试用来测试导入课程功能'''
        course_index = ['课程名', '教师名', '教师电话', '星期',
                        '上课时间', '总课时', '课长', '价格', '开始日期']
        course_data = ['课程名', '测试老师', '测试电话', '星期二',
                       '20:20:00', '30', '45', '70', '2018-8-27']
        course_data2 = ['课程名', '测试老师', '测试电话', '星期三',
                        '20:20:00', '30', '45', '70', '2018-8-27']
        new_file = Workbook()
        sheet = new_file.active
        sheet.append(course_index)
        sheet.append(course_data)
        sheet.append(course_data2)
        new_file.save('test.xlsx')
        result = import_courses('test.xlsx', 'test.xlsx')
        os.remove('test.xlsx')
        self.assertEqual(result[0], STATUS_CODE['Success'])


class CreateAuditionRecordsTestCase(TestCase):
    '''本类用于测试填写试听记录'''

    def setUp(self):
        consult = User.objects.create(
            username='consult', user_type=USER_TYPE['consultant'])
        consult.set_password('consult')
        consult.save()
        user = User.objects.create(
            username='customer', user_type=USER_TYPE['customer'])
        Customer.objects.create(user=user, child_name='child')
        teacher = Teacher.objects.create(name='teacher', tel='13845267894')
        course = Courses.objects.create(
            name='course', teacher_id=teacher, time='8:00:00',
            date=DATE_CODE['Monday'], total_sec=48, price=1024)
        Section.objects.create(
            course_id=course, count=3, name='course<1>',
            start_time='08:00:00', end_time='10:00:00',
            location='classIn105', date='2018-08-15')

    def test_add_nosign_record(self):
        '''测试增添没有报名的试听记录'''
        client = Client()
        client.login(username='consult', password='consult')
        customer = Customer.objects.get(child_name='child')
        data = {
            'id': customer.id,
            'section_name': 'course<1>',
            'is_signedup': False,
            'info': '孩子很喜欢'
        }
        rep = client.post(
            '/data/create_audition_records', data,
            content_type='application/json')
        content = json.loads(rep.content)
        customer = Customer.objects.get(child_name='child')
        self.assertEqual(rep.status_code, 200)
        self.assertEqual(content['error'], STATUS_CODE['Success'])
        self.assertEqual(customer.is_signedup, False)

    def test_add_sign_record(self):
        '''测试增添报名的试听记录'''
        client = Client()
        client.login(username='consult', password='consult')
        customer = Customer.objects.get(child_name='child')
        data = {
            'id': customer.id,
            'section_name': 'course<1>',
            'is_signedup': True,
            'info': '孩子很喜欢'
        }
        rep = client.post(
            '/data/create_audition_records', data,
            content_type='application/json')
        content = json.loads(rep.content)
        customer = Customer.objects.get(child_name='child')
        self.assertEqual(rep.status_code, 200)
        self.assertEqual(content['error'], STATUS_CODE['Success'])
        self.assertEqual(customer.audition_count, 1)

    def test_nosection_record(self):
        '''测试输入的课节名称不存在'''
        client = Client()
        client.login(username='consult', password='consult')
        customer = Customer.objects.get(child_name='child')
        data = {
            'id': customer.id,
            'section_name': 'course<no>',
            'is_signedup': True,
            'info': '孩子很喜欢'
        }
        rep = client.post(
            '/data/create_audition_records', data,
            content_type='application/json')
        content = json.loads(rep.content)
        customer = Customer.objects.filter()
        self.assertEqual(rep.status_code, 200)
        self.assertEqual(content['error'], STATUS_CODE['Not Found'])

    def test_nocustomer_record(self):
        '''测试输入的用户id不存在'''
        client = Client()
        client.login(username='consult', password='consult')
        customer = Customer.objects.get(child_name='child')
        data = {
            'id': customer.id+1,
            'section_name': 'course<1>',
            'is_signedup': True,
            'info': '孩子很喜欢'
        }
        rep = client.post(
            '/data/create_audition_records', data,
            content_type='application/json')
        content = json.loads(rep.content)
        self.assertEqual(rep.status_code, 200)
        self.assertEqual(content['error'], STATUS_CODE['User Not Found'])
